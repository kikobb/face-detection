import sys
from openpyxl import Workbook, load_workbook


def parse_arg():
    if len(sys.argv) != 4:
        exit(1)
    return sys.argv[1], sys.argv[2], sys.argv[3]


# def iter_rows(ws, n):  # produce the list of items in the particular row
#     for row in ws.iter_rows(n):
#         yield [cell.value for cell in row]


def main():
    wb1_path, wb2_path, name = parse_arg()
    wb1 = load_workbook(filename=wb1_path)
    wb2 = load_workbook(filename=wb2_path)

    ws2 = wb2.active

    # for row in ws2.iter_rows():
    #     for cell in row:
    #         if cell.value == 'TrueValue':
    #             n = 'A' + str(cell.row) + ':' + ('GH' + str(cell.row))
    #             list_to_append = list(iter_rows(ws2, n))
    #             for items in list_to_append:
    #                 wb1.active.append(items)

    for row in ws2.iter_rows(min_row=ws2.min_row + 1, max_row=ws2.max_row):
        row_to_append = []
        for cell in row:
            row_to_append.append(cell.value)
        wb1.active.append(row_to_append)

    wb1.save(filename=name)


if __name__ == '__main__':
    # cProfile.run(main(sys.argv[1:]))
    main()
