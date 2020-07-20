import sys
from openpyxl import Workbook, load_workbook
from os import path


def parse_arg():
    if len(sys.argv) == 1 or len(sys.argv) > 3:
        exit(1)
    # append to default output
    if len(sys.argv) == 2:
        return 'compl_exp_1.xlsx', sys.argv[1]
    # append to specific output
    return sys.argv[1], sys.argv[2]


# def iter_rows(ws, n):  # produce the list of items in the particular row
#     for row in ws.iter_rows(n):
#         yield [cell.value for cell in row]


def main():
    wb_out_path, wb_in_path = parse_arg()
    # print(f'in: {wb_in_path}, out: {wb_out_path}')
    margin = 0
    # check if wb_in file exist
    if not path.exists(wb_in_path):
        exit(1)
    else:
        ws_in = load_workbook(filename=wb_in_path).active
    # check if create or append to output file
    if path.exists(wb_out_path):
        wb_out = load_workbook(filename=wb_out_path)
        margin = 1
    else:
        wb_out = Workbook()

    ws_out = wb_out.active

    # for row in ws2.iter_rows():
    #     for cell in row:
    #         if cell.value == 'TrueValue':
    #             n = 'A' + str(cell.row) + ':' + ('GH' + str(cell.row))
    #             list_to_append = list(iter_rows(ws2, n))
    #             for items in list_to_append:
    #                 wb1.active.append(items)

    for row in ws_in.iter_rows(min_row=ws_in.min_row + margin, max_row=ws_in.max_row):
        row_to_append = []
        for cell in row:
            row_to_append.append(cell.value)
        ws_out.append(row_to_append)

    wb_out.save(filename=wb_out_path)
    wb_out.close()
    wb_in.close()

if __name__ == '__main__':
    # cProfile.run(main(sys.argv[1:]))
    main()
