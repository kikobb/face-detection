from openpyxl import Workbook, load_workbook


def main():
    base_folder = '/home/k/PycharmProjects/face-detection/experiment_1/data'
    wb_in = load_workbook(f'{base_folder}/normalized_res_exp.xlsx')

    wb_out = Workbook()
    wb_out_sheets = [wb_out.create_sheet("582"), wb_out.create_sheet("509"), wb_out.create_sheet("300"),
                     wb_out.create_sheet("221"), wb_out.create_sheet("209"), wb_out.create_sheet("154"),
                     wb_out.create_sheet("153"), wb_out.create_sheet("107"), wb_out.create_sheet("99"),
                     wb_out.create_sheet("97"), wb_out.create_sheet("71"), wb_out.create_sheet("69"),
                     wb_out.create_sheet("59"), wb_out.create_sheet("56"), wb_out.create_sheet("50"),
                     wb_out.create_sheet("43"), wb_out.create_sheet("39"), wb_out.create_sheet("32"),
                     wb_out.create_sheet("30"), wb_out.create_sheet("20"), wb_out.create_sheet("18"),
                     wb_out.create_sheet("11")]

    # process input data
    meas_per_diff = 300  # number of measurements per difficulty
    in_sheets = ("pc cpu", "pc gpu", "pc myriad", "Nvidia", "Pi CPU", "Pi  MYRIAD")
    for out_offset, sheet in enumerate(in_sheets):
        if sheet == "Pi CPU":
            dev_name = "Raspberry Pi 4: CPU"
        elif sheet == "Pi  MYRIAD":
            dev_name = "Raspberry Pi 4: MYRIAD"
        else:
            dev_name = wb_in[sheet]['A2'].value
        for in_offset, difficulty in enumerate(range(22)):
            if in_offset == 0:  # create header
                wb_out_sheets[difficulty]['A1'] = 'Device name'
                wb_out_sheets[difficulty]['B1'] = 'Individual_inference_execution (μs)'
            for measurement in range(1, meas_per_diff + 1):     # index starts from 1 (as it is in excel)
                wb_out_sheets[difficulty].cell(row=out_offset * meas_per_diff + measurement + 1,
                                               column=1,
                                               value=dev_name)
                wb_out_sheets[difficulty].cell(row=out_offset * meas_per_diff + measurement + 1,
                                               column=2,
                                               value=wb_in[sheet].cell(row=in_offset * meas_per_diff + measurement + 1,
                                                                       column=10).value)
    wb_in.close()
    wb_in = load_workbook(f'{base_folder}/RP3_CPU_res_exp_1.xlsx')
    for in_offset, difficulty in enumerate(range(22)):
        for measurement in range(1, meas_per_diff + 1):  # index starts from 1 (as it is in excel)
            wb_out_sheets[difficulty].cell(row=len(in_sheets) * meas_per_diff + measurement + 1,
                                           column=1,
                                           value='Raspberry Pi 3: CPU')
            wb_out_sheets[difficulty].cell(row=len(in_sheets) * meas_per_diff + measurement + 1,
                                           column=2,
                                           value=wb_in.active.cell(row=in_offset * meas_per_diff + measurement + 1,
                                                                   column=10).value)
    wb_in.close()
    wb_in = load_workbook(f'{base_folder}/RP3_MYRIAD_res_exp_1.xlsx')
    for in_offset, difficulty in enumerate(range(22)):
        for measurement in range(1, meas_per_diff + 1):  # index starts from 1 (as it is in excel)
            wb_out_sheets[difficulty].cell(row=(len(in_sheets)+1) * meas_per_diff + measurement + 1,
                                           column=1,
                                           value='Raspberry Pi 3: MYRIAD')
            wb_out_sheets[difficulty].cell(row=(len(in_sheets)+1) * meas_per_diff + measurement + 1,
                                           column=2,
                                           value=wb_in.active.cell(row=in_offset * meas_per_diff + measurement + 1,
                                                                   column=10).value)
    wb_in.close()
    wb_out.remove(wb_out.get_sheet_by_name('Sheet'))
    wb_out.save(f'{base_folder}/difficulty_sorted_data.xlsx')
    wb_out.close()


if __name__ == '__main__':
    # cProfile.run(main(sys.argv[1:]))
    main()
